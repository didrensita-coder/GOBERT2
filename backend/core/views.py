# core/views.py
from django.views.decorators.csrf import csrf_exempt
from rest_framework.parsers import MultiPartParser, FormParser
from django.utils.decorators import method_decorator
from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework.views import APIView
from django.contrib.auth import login, logout
from django.http import HttpResponse
from .models import Usuario, Equipo, Accion, Departamento
from .serializers import (
    UsuarioSerializer, RegistroUsuarioSerializer, 
    LoginSerializer, EquipoSerializer, AccionSerializer, DepartamentoSerializer
)
import csv

# ============================================
# FUNCIÓN PARA REGISTRAR ACCIONES
# ============================================

def registrar_accion(usuario, tipo, entidad, entidad_id, descripcion, request=None):
    """Función auxiliar para registrar acciones"""
    ip = None
    if request:
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
    
    Accion.objects.create(
        usuario=usuario,
        tipo=tipo,
        entidad=entidad,
        entidad_id=entidad_id,
        descripcion=descripcion,
        ip=ip
    )


# ============================================
# PERMISOS PERSONALIZADOS
# ============================================

class IsAdminUser(permissions.BasePermission):
    """Permiso solo para administradores"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.rol == 'admin'


# ============================================
# VISTAS DE AUTENTICACIÓN
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class LoginView(APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        print("=== INTENTO DE LOGIN ===")
        print(f"Datos: {request.data}")
        
        serializer = LoginSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.validated_data['user']
            login(request, user)
            print(f"✅ Login exitoso: {user.username}")
            
            # Registrar acción
            registrar_accion(user, 'login', 'perfil', user.id, "Inicio de sesión", request)
            
            user_serializer = UsuarioSerializer(user)
            user_data = user_serializer.data
            
            if user.foto_perfil:
                user_data['foto_perfil'] = request.build_absolute_uri(user.foto_perfil.url)
            else:
                user_data['foto_perfil'] = None
            
            return Response({
                'success': True,
                'user': user_data
            })
        print(f"❌ Error: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@method_decorator(csrf_exempt, name='dispatch')
class LogoutView(APIView):
    def post(self, request):
        if request.user.is_authenticated:
            registrar_accion(request.user, 'logout', 'perfil', request.user.id, "Cierre de sesión", request)
        logout(request)
        return Response({'success': True})


@method_decorator(csrf_exempt, name='dispatch')
class VerificarAuthView(APIView):
    def get(self, request):
        if request.user.is_authenticated:
            user_data = {
                'id': request.user.id,
                'username': request.user.username,
                'rol': request.user.rol,
            }
            if request.user.foto_perfil:
                user_data['foto_perfil'] = request.build_absolute_uri(request.user.foto_perfil.url)
            else:
                user_data['foto_perfil'] = None
            
            return Response({
                'authenticated': True,
                'user': user_data
            })
        return Response({'authenticated': False})


# ============================================
# VISTAS DE USUARIOS
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class UserViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    permission_classes = [IsAdminUser]
    
    @action(detail=False, methods=['post'], permission_classes=[IsAdminUser])
    def registrar_coordinador(self, request):
        """Registrar nuevo coordinador (solo admin)"""
        print("=== REGISTRANDO COORDINADOR ===")
        print(f"Datos: {request.data}")
        
        serializer = RegistroUsuarioSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            print(f"✅ Coordinador creado: {user.username}")
            
            registrar_accion(request.user, 'creacion', 'usuario', user.id, f"Creó el usuario {user.username}", request)
            
            return Response(UsuarioSerializer(user).data, status=status.HTTP_201_CREATED)
        print(f"❌ Error: {serializer.errors}")
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, *args, **kwargs):
        user = self.get_object()
        username = user.username
        registrar_accion(request.user, 'eliminacion', 'usuario', user.id, f"Eliminó el usuario {username}", request)
        return super().destroy(request, *args, **kwargs)
    
    @action(detail=False, methods=['get'])
    def me(self, request):
        """Obtener usuario actual"""
        serializer = UsuarioSerializer(request.user)
        user_data = serializer.data
        if request.user.foto_perfil:
            user_data['foto_perfil'] = request.build_absolute_uri(request.user.foto_perfil.url)
        return Response(user_data)
    
    @action(detail=False, methods=['put'], permission_classes=[permissions.IsAuthenticated])
    def actualizar_perfil(self, request):
        """Actualizar el perfil del usuario autenticado"""
        user = request.user
        
        if 'first_name' in request.data:
            user.first_name = request.data['first_name']
        if 'last_name' in request.data:
            user.last_name = request.data['last_name']
        if 'email' in request.data:
            user.email = request.data['email']
        if 'telefono' in request.data:
            user.telefono = request.data['telefono']
        if 'departamento' in request.data:
            user.departamento = request.data['departamento']
        if 'foto_perfil' in request.FILES:
            if user.foto_perfil:
                user.foto_perfil.delete()
            user.foto_perfil = request.FILES['foto_perfil']
        
        user.save()
        
        registrar_accion(request.user, 'edicion', 'perfil', user.id, "Actualizó su perfil", request)
        
        serializer = UsuarioSerializer(user)
        user_data = serializer.data
        
        if user.foto_perfil:
            user_data['foto_perfil'] = request.build_absolute_uri(user.foto_perfil.url)
        else:
            user_data['foto_perfil'] = None
        
        return Response({
            'success': True,
            'user': user_data
        })
    
    @action(detail=False, methods=['post'], permission_classes=[permissions.IsAuthenticated])
    def cambiar_password(self, request):
        """Cambiar la contraseña del usuario autenticado"""
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')
        
        if not user.check_password(current_password):
            return Response(
                {'errors': {'current_password': 'La contraseña actual es incorrecta'}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if len(new_password) < 6:
            return Response(
                {'errors': {'new_password': 'La contraseña debe tener al menos 6 caracteres'}},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        user.set_password(new_password)
        user.save()
        
        registrar_accion(request.user, 'edicion', 'perfil', user.id, "Cambió su contraseña", request)
        
        return Response({'success': True, 'message': 'Contraseña actualizada correctamente'})


# ============================================
# VISTAS DE EQUIPOS
# ============================================

@method_decorator(csrf_exempt, name='dispatch')
class EquipoViewSet(viewsets.ModelViewSet):
    queryset = Equipo.objects.all()
    serializer_class = EquipoSerializer
    parser_classes = [MultiPartParser, FormParser]
    
    def get_permissions(self):
        return [permissions.AllowAny()]
    
    def perform_create(self, serializer):
        user = self.request.user if self.request.user.is_authenticated else None
        equipo = serializer.save(registrado_por=user)
        if user:
            registrar_accion(user, 'creacion', 'equipo', equipo.id, f"Creó el equipo {equipo.codigo_equipo}", self.request)
    
    def perform_update(self, serializer):
        equipo = serializer.save()
        user = self.request.user if self.request.user.is_authenticated else None
        if user:
            registrar_accion(user, 'edicion', 'equipo', equipo.id, f"Editó el equipo {equipo.codigo_equipo}", self.request)
    
    def perform_destroy(self, instance):
        user = self.request.user if self.request.user.is_authenticated else None
        codigo = instance.codigo_equipo
        if user:
            registrar_accion(user, 'eliminacion', 'equipo', instance.id, f"Eliminó el equipo {codigo}", self.request)
        instance.delete()
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        equipos = Equipo.objects.all()
        stats = {
            'total': equipos.count(),
            'bueno': equipos.filter(estado='bueno').count(),
            'regular': equipos.filter(estado='regular').count(),
            'malo': equipos.filter(estado='malo').count(),
            'por_tipo': {},
            'por_ubicacion': {},
        }
        
        for equipo in equipos:
            tipo = equipo.get_tipo_display()
            stats['por_tipo'][tipo] = stats['por_tipo'].get(tipo, 0) + 1
            ubicacion = equipo.ubicacion
            stats['por_ubicacion'][ubicacion] = stats['por_ubicacion'].get(ubicacion, 0) + 1
        
        return Response(stats)
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def exportar_excel(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Equipos"
        
        headers = ['Código', 'Tipo', 'Ubicación', 'Usuario', 'Procesador', 'RAM', 'Disco', 'SO', 'Estado', 'Fecha']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        for equipo in self.queryset:
            ws.append([
                equipo.codigo_equipo,
                equipo.get_tipo_display(),
                equipo.ubicacion,
                equipo.usuario_asignado,
                equipo.procesador,
                equipo.ram,
                equipo.disco_duro,
                equipo.sistema_operativo,
                equipo.get_estado_display(),
                equipo.fecha_registro.strftime('%Y-%m-%d')
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="equipos.xlsx"'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def exportar_csv(self, request):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="equipos.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Código', 'Tipo', 'Ubicación', 'Usuario', 'Procesador', 'RAM', 'Disco', 'SO', 'Estado', 'Fecha'])
        
        for equipo in self.queryset:
            writer.writerow([
                equipo.codigo_equipo,
                equipo.get_tipo_display(),
                equipo.ubicacion,
                equipo.usuario_asignado,
                equipo.procesador,
                equipo.ram,
                equipo.disco_duro,
                equipo.sistema_operativo,
                equipo.get_estado_display(),
                equipo.fecha_registro.strftime('%Y-%m-%d')
            ])
        
        return response


# ============================================
# VISTAS DE ACCIONES
# ============================================

class AccionViewSet(viewsets.ModelViewSet):
    queryset = Accion.objects.all()
    serializer_class = AccionSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        if self.request.user.rol == 'admin':
            return Accion.objects.all()
        return Accion.objects.filter(usuario=self.request.user)


# ============================================
# VISTAS DE DEPARTAMENTOS (NUEVO)
# ============================================

class DepartamentoViewSet(viewsets.ModelViewSet):
    queryset = Departamento.objects.all()
    serializer_class = DepartamentoSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_permissions(self):
        # Solo administradores pueden crear, editar o eliminar
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [permissions.IsAuthenticated()]
    
    def perform_create(self, serializer):
        depto = serializer.save()
        registrar_accion(
            self.request.user, 
            'creacion', 
            'departamento', 
            depto.id, 
            f"Creó el departamento {depto.nombre}", 
            self.request
        )
    
    def perform_update(self, serializer):
        depto = serializer.save()
        registrar_accion(
            self.request.user, 
            'edicion', 
            'departamento', 
            depto.id, 
            f"Editó el departamento {depto.nombre}", 
            self.request
        )
    
    def perform_destroy(self, instance):
        registrar_accion(
            self.request.user, 
            'eliminacion', 
            'departamento', 
            instance.id, 
            f"Eliminó el departamento {instance.nombre}", 
            self.request
        )
        instance.delete()